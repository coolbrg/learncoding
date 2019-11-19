# Given following XLSX data
# Timestamp             Can You attend?     Names                                 How did you hear  Preferred Timing
# 11/19/2019 16: 23: 54	Yes,  I'll be there	brg, shirisha, swopnil, owen, dipesh	Social Media	3 PM - 6 PM
# 11/19/2019 16: 27: 28	Yes,  I'll be there	pradyu, swati, rakesh, rajesh	Friend	3 PM - 6 PM
# 11/19/2019 16: 28: 59	Yes,  I'll be there	prashu, paroskh, mannu, sunita, rajesh, brg	Website	6 PM - 9 PM
# 11/20/2019 16: 28: 59	Yes,  I'll be there	sunita gurung	Website	6 PM - 9 PM
# 11/21/2019 16: 28: 59	Yes,  I'll be there	harsh, sunita, mannu, budhram	Website	6 PM - 9 PM
# 11/22/2019 16: 28: 59	Yes,  I'll be there	jit bahadur, man kumari, santa maya	Social Media	6 PM - 9 PM
# 11/23/2019 16: 28: 59	Yes,  I'll be there	dhanendra, mandira, indira, mandip	Website	3 PM - 6 PM
# 11/24/2019 16: 28: 59	Sorry, can't make it
# 11/25/2019 16: 28: 59	Yes,  I'll be there	rakesh, chetan jadhav, chetan chavan, anushree, sameer	Social Media	12 PM - 9 PM
# 11/26/2019 16: 28: 59	Yes,  I'll be there	dev gurung	Website	6 PM - 9 PM
# 11/27/2019 16: 28: 59	Yes,  I'll be there	dil bahadur, dillu maya	Website	9 AM - 12 Noon
# 11/28/2019 16: 28: 59	Yes,  I'll be there	samsher, bir bahadur, maili	Social Media	3 PM - 6 PM
# 11/29/2019 16: 28: 59	Yes,  I'll be there	rinu	Website	6 PM - 9 PM
# 11/30/2019 16: 28: 59	Sorry, can't make it
# 12/1/2019 16: 28: 59	Yes,  I'll be there	manoj, vaibhav, shailesh	Website	6 PM - 9 PM
# 12/2/2019 16: 28: 59	Sorry, can't make it
# 12/3/2019 16: 28: 59	Yes,  I'll be there	aakash, sudhir, priyansh	Website	9 AM - 12 Noon
# 12/4/2019 16: 28: 59	Sorry, can't make it
# 12/5/2019 16: 28: 59	Yes,  I'll be there	gagan, aakash, neer	Social Media	6 PM - 9 PM

# Link: https://medium.com/aubergine-solutions/working-with-excel-sheets-in-python-using-openpyxl-4f9fd32de87f

from openpyxl import load_workbook

wb = load_workbook(filename='assets/rsvp.xlsx')
print(type(wb))

sheet = wb.active
max_column = sheet.max_column
max_row = sheet.max_row

print("\n\nTotal Rows: {}".format(max_row))
print("Total Columns: {}".format(max_column))

# iterate over all cells and rows
for row in range(1, max_row + 1):
  for col in range(1, max_column + 1):
    cell = sheet.cell(row=row, column=col)
    print(cell.value, end=' | ')

  print("\n")

# print row headers
for col in range(1, max_column + 1):
  cell = sheet.cell(row=1, column=col)
  print(cell.value, end=' | ')

# print total number of attendees and their names
total_attendees = 0
attendee_names = []
for row in range(2, max_row + 1):
  cell = sheet.cell(row=row, column=3)
  if cell.value != None:
    attendees = cell.value.split(",")
    total_attendees += len(attendees)
    attendee_names = attendee_names + attendees

print("Total Attendees: {}".format(total_attendees))
print("Attendees are: {}".format(", ".join(attendee_names)))

